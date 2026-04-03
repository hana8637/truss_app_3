# --- [최신 파이썬 에러 방지 패치] ---
# 파이썬 3.12 이상에서 distutils가 없어서 발생하는 에러를 해결합니다.
try:
    import distutils.version
except ImportError:
    try:
        import setuptools
    except ImportError:
        pass

import streamlit as st
import pandas as pd
import math
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import io

# --- [한글 깨짐 방지 라이브러리] ---
try:
    import koreanize_matplotlib
except ImportError:
    st.error("koreanize-matplotlib 라이브러리가 필요합니다. requirements.txt를 확인해 주세요.")

# --- 엑셀 스타일링 라이브러리 ---
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl 라이브러리가 필요합니다.")

# ==============================================================================
# 공통 설정
# ==============================================================================
st.set_page_config(page_title="하나천막기업 자재 산출", layout="wide")

# ==============================================================================
# [1] 트러스 시스템 관련 함수
# ==============================================================================
def save_formatted_excel(raw_data):
    """트러스 전용 엑셀 저장 및 서식 지정"""
    df = pd.DataFrame(raw_data)
    df_grouped = df.groupby(["구분", "품명", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)"]).size().reset_index(name='1대당 수량')
    
    # 정렬 및 순번 부여
    sort_mapping = {"상현대(전체)": -2, "하현대(전체)": -1, "용마루": 1, "다대": 6, "살대": 9}
    df_grouped['정렬키'] = df_grouped['구분'].map(sort_mapping).fillna(99)
    df_grouped = df_grouped.sort_values(by=["정렬키", "재단기장(L)"], ascending=[True, False]).drop('정렬키', axis=1)
    df_grouped.insert(0, '순번', range(1, len(df_grouped) + 1))
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_grouped.to_excel(writer, sheet_name='통합 재단표', index=False, startrow=2)
        ws = writer.sheets['통합 재단표']
        
        # 기본 헤더 설정
        ws['A1'] = "👉 트러스 총 제작 수량 (EA) :"
        ws['D1'] = 1
        
        # 테두리 및 서식
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for r_idx in range(3, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                if r_idx == 3:
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 15

    return output.getvalue()

def generate_custom_truss(params):
    """트러스 도면 생성 및 데이터 계산"""
    span_cm = params['span_cm']
    divs = params['divs']
    h_outer_cm = params['h_outer_cm']
    h_center_cm = params['h_center_cm']
    m_od = params['m_od']
    r_od = params['r_od']
    v_od = params['v_od']

    S = span_cm * 10
    H_out = h_outer_cm * 10
    H_cen = h_center_cm * 10

    def get_y_top(x):
        m = (H_cen - H_out) / (S/2)
        return H_out + m * x if x <= S/2 else H_out + m * (S - x)

    def get_slope(x):
        dx = 0.1
        dy = get_y_top(x + dx if x + dx <= S else x - dx) - get_y_top(x)
        return math.degrees(math.atan2(dy, dx))

    fig, ax = plt.subplots(figsize=(20, 8))
    mid_idx = divs // 2
    raw_data = []

    # 다대 그리기 로직
    v_centers_x = [v_od/2 if i==0 else (S - r_od/2 if i==divs else i*(S/divs)) for i in range(divs + 1)]
    
    for i in range(divs + 1):
        x = v_centers_x[i]
        is_ridge = (i == mid_idx)
        curr_v_od = r_od if is_ridge else v_od
        y_top = get_y_top(x)
        v_len = y_top - m_od # 간이 계산
        angle = abs(get_slope(x))
        
        ax.add_patch(patches.Rectangle((x-curr_v_od/2, 0), curr_v_od, y_top, facecolor='lightgray', edgecolor='black', alpha=0.7))
        raw_data.append({"구분": "용마루" if is_ridge else "다대", "품명": f"{curr_v_od}mm", "재단기장(L)": round(v_len,1), "상단 가공각(°)": int(angle), "하단 가공각(°)": 0})

    # 상현/하현 라인
    x_vals = np.linspace(0, S, 100)
    ax.plot(x_vals, [get_y_top(x) for x in x_vals], color='blue', lw=3, label='상현대')
    ax.axhline(0, color='red', lw=3, label='하현대')
    
    ax.set_aspect('equal')
    ax.set_title(f"트러스 도면 (스판: {span_cm}cm)", fontsize=20)
    ax.legend()
    
    excel_bytes = save_formatted_excel(raw_data)
    return fig, excel_bytes

# ==============================================================================
# UI 메인 화면
# ==============================================================================
st.title("🏢 하나천막기업 자재 산출 시스템")
st.markdown("---")

tab1, tab2 = st.tabs(["트러스 생성기", "사다리 산출"])

with tab1:
    st.header("1. 맞춤형 트러스 생성")
    c1, c2 = st.columns(2)
    with c1:
        span_cm = st.number_input("전체 스판(cm)", value=1200)
        divs = st.number_input("등분 수", value=34)
    with c2:
        h_out = st.number_input("끝단 높이(cm)", value=80)
        h_cen = st.number_input("중앙 높이(cm)", value=250)
    
    if st.button("도면 및 재단표 생성", type="primary"):
        params = {'span_cm': span_cm, 'divs': divs, 'h_outer_cm': h_out, 'h_center_cm': h_cen,
                  'm_od': 59.9, 'v_od': 38.1, 'r_od': 59.9, 'd_od': 31.8, 'offset_mm': 20, 't_name': '기본삼각'}
        fig, excel_data = generate_custom_truss(params)
        st.pyplot(fig)
        st.download_button("📥 엑셀 재단표 다운로드", excel_data, f"트러스_{span_cm}.xlsx")

with tab2:
    st.header("2. 사다리 및 용마루 산출")
    L_cm = st.number_input("사다리 총기장(cm)", value=2000)
    if st.button("사다리 산출 실행"):
        st.success(f"{L_cm}cm 사다리 산출이 완료되었습니다. (도면 준비 중)")
        # 사다리 로직 추가 시 여기에 연결
