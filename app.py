import streamlit as st
import pandas as pd
import math
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.backends.backend_pdf import PdfPages
import io

# --- 마법의 한글 폰트 자동 설정 라이브러리 ---
# 이 한 줄만 있으면 Streamlit 웹에서도 폰트 파일 없이 한글이 절대 깨지지 않습니다.
try:
    import koreanize_matplotlib
except ImportError:
    st.error("koreanize-matplotlib 라이브러리가 필요합니다. requirements.txt에 추가해 주세요.")

# --- 엑셀 스타일링 라이브러리 ---
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl 라이브러리가 필요합니다. (pip install openpyxl)")

# ==============================================================================
# 공통 설정
# ==============================================================================
st.set_page_config(page_title="하나천막기업 자재 산출", layout="wide")

# (기존의 복잡했던 set_korean_font() 함수는 koreanize_matplotlib가 알아서 다 해주므로 삭제했습니다!)

# ==============================================================================
# [1] 트러스 시스템 관련 함수 
# ==============================================================================
def save_formatted_excel(raw_data, filename):
    """트러스 전용 엑셀 저장 및 서식 지정 로직"""
    df = pd.DataFrame(raw_data)
    df_grouped = df.groupby(["구분", "품명", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)"]).size().reset_index(name='1대당 수량')
    
    sort_mapping = {
        "상현대(전체)": -2, "하현대(전체)": -1,  
        "용마루": 1, "상단용마루": 2, "하단용마루": 3,
        "수평재": 4, "밑더블수평재": 5, 
        "다대": 6, "상단다대": 7, "하단다대": 8,
        "살대": 9, "상단살대": 10, "하단살대": 11,
        "수평내부다대": 12, "수평내부살대": 13, "서브다대": 14, "서브살대": 15
    }
    df_grouped['정렬키'] = df_grouped['구분'].map(sort_mapping).fillna(99)
    df_grouped = df_grouped.sort_values(by=["정렬키", "재단기장(L)"], ascending=[True, False]).drop('정렬키', axis=1)
    
    df_grouped.insert(0, '순번', range(1, len(df_grouped) + 1))
    df_grouped["총 소요 수량"] = ""
    df_grouped["6M 소요본수"] = ""
    df_grouped = df_grouped[["순번", "구분", "품명", "1대당 수량", "총 소요 수량", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)", "6M 소요본수"]]

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_grouped.to_excel(writer, sheet_name='통합 재단표', index=False, startrow=2)
        ws = writer.sheets['통합 재단표']
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "👉 트러스 총 제작 수량 (EA) :"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal="right", vertical="center")
        
        ws['D1'] = 1
        ws['D1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['D1'].font = Font(color="FF0000", bold=True, size=14)
        ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
                             top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))
        ws['D1'].border = thin_border

        color_map = {
            "상현대(전체)": PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid"),
            "하현대(전체)": PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid"),
            "용마루": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "상단용마루": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
            "하단용마루": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
            "수평재": PatternFill(start_color="D2B4DE", end_color="D2B4DE", fill_type="solid"), 
            "밑더블수평재": PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"),
            "다대": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            "상단다대": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
            "하단다대": PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid"),
            "살대": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "상단살대": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
            "하단살대": PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid"),
            "수평내부다대": PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid"), 
            "수평내부살대": PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid"), 
            "서브다대": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
            "서브살대": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        }
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 3):
            gubun_val = ws.cell(row=r_idx, column=2).value if r_idx > 3 else None
            for c_idx, cell in enumerate(row, 1):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
                if r_idx == 3:
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    header = ws.cell(row=3, column=c_idx).value
                    if header == "순번": cell.font = Font(bold=True)
                    elif header == "구분":
                        cell.fill = color_map.get(gubun_val, PatternFill(fill_type=None))
                        cell.font = Font(bold=True)
                    elif header == "총 소요 수량":
                        cell.value = f'=$D$1*D{r_idx}' 
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        cell.font = Font(color="0070C0", bold=True)
                    elif header == "6M 소요본수":
                        cell.value = f'=ROUNDUP((E{r_idx}*F{r_idx})/6000, 1)'
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        cell.font = Font(color="0070C0", bold=True)
                    elif header in ["1대당 수량", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)"]:
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="C00000", bold=True)
                    elif r_idx % 2 == 0: cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        start_col = 11  
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+4)
        title_cell = ws.cell(row=1, column=start_col, value="레이저 가공 싸이즈")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        title_cell.border = thin_border
        
        headers_new = ["구분", "재단기장\n(반올림)", "상단 가공각", "하단 가공각", "총수량"]
        for i, h in enumerate(headers_new):
            c = ws.cell(row=3, column=start_col+i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border

        counters = {}
        row_idx_new = 4
        for _, row_data in df_grouped.iterrows():
            cat_base = row_data["구분"]
            if cat_base not in counters: counters[cat_base] = 1
            name_val = f"{cat_base}{counters[cat_base]}"
            counters[cat_base] += 1
            
            L_val = row_data["재단기장(L)"]
            top_a = row_data["상단 가공각(°)"]
            bot_a = row_data["하단 가공각(°)"]
            qty_val = row_data["1대당 수량"]
            
            L_rounded = round(float(L_val)) if pd.notnull(L_val) else 0
            top_ceil = int(top_a) if pd.notnull(top_a) else 0
            bot_ceil = int(bot_a) if pd.notnull(bot_a) else 0
            
            ws.cell(row=row_idx_new, column=start_col, value=name_val)
            ws.cell(row=row_idx_new, column=start_col+1, value=L_rounded)
            ws.cell(row=row_idx_new, column=start_col+2, value=top_ceil)
            ws.cell(row=row_idx_new, column=start_col+3, value=bot_ceil)
            ws.cell(row=row_idx_new, column=start_col+4, value=f"=$D$1*{qty_val}") 
            ws.cell(row=row_idx_new, column=start_col+4).font = Font(color="0070C0", bold=True)
            
            for i in range(5):
                c = ws.cell(row=row_idx_new, column=start_col+i)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
                if row_idx_new % 2 == 0: 
                    c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            row_idx_new += 1

        pipe_summary = {}
        for item in raw_data:
            p_name = item["품명"]
            l_val = float(item["재단기장(L)"]) if pd.notnull(item["재단기장(L)"]) else 0
            if p_name not in pipe_summary:
                pipe_summary[p_name] = {"len": 0, "qty": 0}
            pipe_summary[p_name]["len"] += l_val
            pipe_summary[p_name]["qty"] += 1
            
        sum_start_col = 17 
        ws.merge_cells(start_row=1, start_column=sum_start_col, end_row=1, end_column=sum_start_col+3)
        sum_title = ws.cell(row=1, column=sum_start_col, value="📦 파이프 규격별 발주/재단 총괄표")
        sum_title.font = Font(bold=True, size=14)
        sum_title.alignment = Alignment(horizontal="center", vertical="center")
        sum_title.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        sum_title.border = thin_border
        
        sum_headers = ["파이프 규격", "총 절단 수량(EA)", "총 소요길이(mm)", "6M 발주(본)"]
        for i, h in enumerate(sum_headers):
            c = ws.cell(row=3, column=sum_start_col+i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="385623", end_color="385623", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border
            
        sum_r_idx = 4
        for p_name, data in pipe_summary.items():
            total_l = round(data["len"], 1)
            base_qty = data["qty"]
            
            c1 = ws.cell(row=sum_r_idx, column=sum_start_col, value=p_name)
            c2 = ws.cell(row=sum_r_idx, column=sum_start_col+1, value=f"={base_qty}*$D$1")
            c3 = ws.cell(row=sum_r_idx, column=sum_start_col+2, value=f"={total_l}*$D$1")
            c4 = ws.cell(row=sum_r_idx, column=sum_start_col+3, value=f"=ROUNDUP(({total_l}*$D$1)/6000, 0)")
            
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
            c3.font = Font(color="0070C0", bold=True)
            c4.font = Font(color="FF0000", bold=True, size=12)
            
            for i in range(4):
                c = ws.cell(row=sum_r_idx, column=sum_start_col+i)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
                if sum_r_idx % 2 == 0:
                    c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            sum_r_idx += 1

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max([len(str(ws.cell(row=r, column=col_idx).value)) for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=col_idx).value] + [0])
            ws.column_dimensions[col_letter].width = max(max_len + 6, 12)

def generate_custom_truss(params):
    """트러스 생성 메인 함수 (Streamlit 파라미터 매핑)"""
    t_name = params['t_name']
    type_choice = params['type_choice']
    span_cm = params['span_cm']
    divs = params['divs']
    h_outer_cm = params['h_outer_cm']
    h_center_cm = params['h_center_cm']
    h_tie_cm = params['h_tie_cm']
    m_od = params['m_od']
    v_od = params['v_od']
    r_od = params['r_od']
    d_od = params['d_od']
    offset_mm = params['offset_mm']
    
    is_sub_type = type_choice in ["4", "5", "6"]
    has_tie = type_choice in ["4", "5"] 
    is_double_bot = type_choice in ["7", "8", "9"]
    is_half = type_choice in ["3", "6", "9"]
    is_arch = type_choice in ["2", "5", "8"]

    S, H_out, H_cen, H_tie = span_cm * 10, h_outer_cm * 10, h_center_cm * 10, h_tie_cm * 10

    yc, R = 0, 0
    if is_arch:
        if H_cen <= H_out: H_cen = H_out + 10
        yc = ((S/2)**2 + H_out**2 - H_cen**2) / (2 * (H_out - H_cen))
        R = H_cen - yc

    def get_y_top(x):
        if x < 0: x = 0
        if x > S: x = S
        if type_choice in ["1", "4", "7"]:
            m = (H_cen - H_out) / (S/2)
            return H_out + m * x if x <= S/2 else H_out + m * (S - x)
        elif is_arch:
            val = max(R**2 - (x - S/2)**2, 0)
            return yc + math.sqrt(val)
        elif is_half:
            return H_out + (x / S) * (H_cen - H_out)

    def get_y_bot(x):
        if is_sub_type: return max(get_y_top(x) - H_out, 0.0)
        else: return 0.0

    def get_slope(func, x):
        dx = 0.1
        test_x = x if x + dx <= S else x - dx
        dy = func(test_x + dx) - func(test_x)
        return math.degrees(math.atan2(dy, dx))
        
    def get_cos(func, x):
        dx = 0.1
        test_x = x if x + dx <= S else x - dx
        dy = func(test_x + dx) - func(test_x)
        cos_val = math.cos(math.atan2(dy, dx))
        return cos_val if cos_val != 0 else 0.0001
        
    def get_thick(func, x, od):
        return od / get_cos(func, x)

    def get_chord_y_top(x):
        return get_y_top(x) - get_thick(get_y_top, x, m_od)
        
    def get_chord_y_bot(x):
        return get_y_bot(x) + get_thick(get_y_bot, x, m_od)

    def draw_dim_text(ax, x, y, text, angle=0, color='black', fontsize=11.5):
        if angle > 90: angle -= 180
        elif angle < -90: angle += 180
        ax.text(x, y, text, color=color, fontsize=fontsize, fontweight='bold', ha='center', va='center', rotation=angle,
                bbox=dict(facecolor='white', alpha=0.85, edgecolor='none', pad=1.5))

    fig, ax = plt.subplots(figsize=(40, 18), dpi=100)
    mid_idx = divs if is_half else divs // 2
    raw_data = []

    v_centers_x = [v_od/2 if i==0 else (S - (v_od if is_half else r_od)/2 if i==divs else i*(S/divs)) for i in range(divs + 1)]
    chord_x = [0] + [v_centers_x[i] for i in range(1, divs)] + [S]

    end_thick = get_thick(get_y_top, 0, m_od)
    H_mid_top = H_out - end_thick if is_double_bot else H_out
    H_mid_bot = H_mid_top - m_od if is_double_bot else H_out - m_od

    total_top_chord_len = sum(math.hypot(chord_x[i+1] - chord_x[i], get_y_top(chord_x[i+1]) - get_y_top(chord_x[i])) for i in range(divs))
    total_bot_chord_len = sum(math.hypot(chord_x[i+1] - chord_x[i], get_y_bot(chord_x[i+1]) - get_y_bot(chord_x[i])) for i in range(divs))
    
    raw_data.append({
        "구분": "상현대(전체)", "품명": f"{m_od}mm 파이프",
        "재단기장(L)": round(total_top_chord_len, 1), "상단 가공각(°)": 0, "하단 가공각(°)": 0
    })
    raw_data.append({
        "구분": "하현대(전체)", "품명": f"{m_od}mm 파이프",
        "재단기장(L)": round(total_bot_chord_len, 1), "상단 가공각(°)": 0, "하단 가공각(°)": 0
    })

    if is_double_bot:
        total_mid_pipe_len = S
        ax.add_patch(patches.Rectangle((0, H_mid_bot), total_mid_pipe_len, m_od, facecolor='#9b59b6', edgecolor='black', zorder=6))
        draw_dim_text(ax, S/2, H_mid_bot + m_od/2, f"밑더블 수평재(전체) L:{total_mid_pipe_len:.1f}", angle=0, color='purple', fontsize=14)
        raw_data.append({
            "구분": "밑더블수평재", "품명": f"{m_od}mm 파이프",
            "재단기장(L)": round(total_mid_pipe_len, 1), "상단 가공각(°)": 0, "하단 가공각(°)": 0
        })

    for i in range(divs + 1):
        x = v_centers_x[i]
        is_ridge = (i == mid_idx) and not is_half
        curr_v_od = r_od if is_ridge else v_od
        
        x_l, x_r = x - curr_v_od/2, x + curr_v_od/2
        
        yt_l = get_chord_y_top(x_l)
        yt_r = get_chord_y_top(x_r)
        yb_l = get_chord_y_bot(x_l)
        yb_r = get_chord_y_bot(x_r)
        
        y_bot_c = get_chord_y_bot(x)
        y_top_c = get_chord_y_top(x)
        
        if x_l < S/2 < x_r and not is_arch:
            poly_top = [[x_r, yt_r], [S/2, get_chord_y_top(S/2)], [x_l, yt_l]]
            poly_bot = [[x_l, yb_l], [S/2, get_chord_y_bot(S/2)], [x_r, yb_r]]
        else:
            poly_top = [[x_r, yt_r], [x_l, yt_l]]
            poly_bot = [[x_l, yb_l], [x_r, yb_r]]
        
        t_angle = int(round(abs(get_slope(get_y_top, x))))
        b_angle = int(round(abs(get_slope(get_y_bot, x))))
        
        if is_ridge: v_cut_l = y_top_c - min(yb_l, yb_r)
        else: v_cut_l = max(yt_l, yt_r) - min(yb_l, yb_r) 

        if has_tie and is_ridge and H_tie > 0:
            if type_choice != "5":
                yb_l = yb_r = y_bot_c = H_tie + m_od/2
                poly_bot = [[x_l, yb_l], [x_r, yb_r]]
                v_cut_l = y_top_c - y_bot_c
                b_angle = 0

        if is_double_bot:
            u_len = y_top_c - H_mid_top
            if u_len > 0:
                u_cut_max = y_top_c - H_mid_top if is_ridge else max(yt_l, yt_r) - H_mid_top
                pts_u = [[x_l, H_mid_top], [x_r, H_mid_top]] + poly_top
                ax.add_patch(patches.Polygon(pts_u, facecolor='#2980b9', edgecolor='black', zorder=5))
                
                g_name = "상단용마루" if is_ridge else "상단다대"
                raw_data.append({
                    "구분": g_name, "품명": f"{curr_v_od}mm 파이프",
                    "재단기장(L)": round(u_cut_max, 1), "상단 가공각(°)": t_angle, "하단 가공각(°)": 0
                })
                stagger_top = 600 if i % 2 == 0 else 900
                my_top = y_top_c + stagger_top
                ax.plot([x, x], [y_top_c + m_od/2, my_top - 180], color='blue', linestyle=':', lw=1.5, zorder=1)
                draw_dim_text(ax, x, my_top, f"{g_name}\nL:{u_cut_max:.1f}", angle=90, color='blue', fontsize=10)

            l_len = H_mid_bot - y_bot_c
            if l_len > 0:
                l_cut_max = H_mid_bot - min(yb_l, yb_r) if not is_ridge else H_mid_bot - y_bot_c
                pts_l = poly_bot + [[x_r, H_mid_bot], [x_l, H_mid_bot]]
                ax.add_patch(patches.Polygon(pts_l, facecolor='#34495e', edgecolor='black', zorder=5))
                
                g_name = "하단용마루" if is_ridge else "하단다대"
                raw_data.append({
                    "구분": g_name, "품명": f"{curr_v_od}mm 파이프",
                    "재단기장(L)": round(l_cut_max, 1), "상단 가공각(°)": 0, "하단 가공각(°)": b_angle
                })
                stagger_bot = 600 if i % 2 == 0 else 900
                my_bot = y_bot_c - stagger_bot
                ax.plot([x, x], [y_bot_c - m_od/2, my_bot + 180], color='darkblue', linestyle=':', lw=1.5, zorder=1)
                draw_dim_text(ax, x, my_bot, f"{g_name}\nL:{l_cut_max:.1f}", angle=90, color='darkblue', fontsize=10)
        else:
            pts_v = poly_bot + poly_top
            ax.add_patch(patches.Polygon(pts_v, facecolor='#2c3e50', edgecolor='black', zorder=5))
            
            text_color = 'red' if is_ridge else 'blue'
            stagger_offset = 600 if i % 2 == 0 else 900
            my = y_top_c + stagger_offset
            ax.plot([x, x], [y_top_c + m_od/2, my - 180], color=text_color, linestyle=':', lw=1.5, zorder=1)
            draw_dim_text(ax, x, my, f"L:{v_cut_l:.1f} (상:{t_angle}°/하:{b_angle}°)", angle=90, color=text_color)
            
            v_gubun = "용마루" if is_ridge else "다대"
            raw_data.append({
                "구분": v_gubun, "품명": f"{curr_v_od}mm 파이프",
                "재단기장(L)": round(v_cut_l, 1), "상단 가공각(°)": t_angle, "하단 가공각(°)": b_angle
            })

    is_diag = True 
    for i in range(divs):
        cx, cnx = chord_x[i], chord_x[i+1]
        
        pb1, pb2 = (cx, get_y_bot(cx)), (cnx, get_y_bot(cnx))
        pb3, pb4 = (cnx, get_chord_y_bot(cnx)), (cx, get_chord_y_bot(cx))
        ax.add_patch(patches.Polygon(np.array([pb1, pb2, pb3, pb4]), facecolor='#7f8c8d', alpha=0.5, zorder=2))
        
        pt1, pt2 = (cx, get_chord_y_top(cx)), (cnx, get_chord_y_top(cnx))
        pt3, pt4 = (cnx, get_y_top(cnx)), (cx, get_y_top(cx))
        ax.add_patch(patches.Polygon(np.array([pt1, pt2, pt3, pt4]), facecolor='#7f8c8d', zorder=7))

        if is_diag:
            x, nx = v_centers_x[i], v_centers_x[i+1]
            
            is_r_curr = (i == mid_idx) and not is_half
            is_r_next = (i+1 == mid_idx) and not is_half
            c_v_od, n_v_od = (r_od if is_r_curr else v_od), (r_od if is_r_next else v_od)

            wx_start = x + c_v_od/2 + offset_mm
            wx_end = nx - n_v_od/2 - offset_mm
            
            def draw_diag(left_x, right_x, is_forward):
                mid_x = (left_x + right_x) / 2
                y_top_est = get_chord_y_top(mid_x)
                y_bot_est = get_chord_y_bot(mid_x)
                v_len = abs(y_top_est - y_bot_est)
                dx = right_x - left_x
                if dx <= 0: return
                diag_len = math.hypot(dx, v_len)
                sin_theta = v_len / diag_len if diag_len > 0 else 1
                w_half = (d_od / 2) / sin_theta if sin_theta > 0.01 else d_od / 2
                
                px_bot = left_x + w_half if is_forward else right_x - w_half
                px_top = right_x - w_half if is_forward else left_x + w_half
                
                py_bot = get_chord_y_bot(px_bot)
                py_top = get_chord_y_top(px_top)
                
                diag_l = math.hypot(px_top - px_bot, py_top - py_bot)
                
                if is_forward:
                    x_bl, x_br = px_bot - w_half, px_bot + w_half
                    x_tr, x_tl = px_top + w_half, px_top - w_half
                    pts = [
                        [x_bl, get_chord_y_bot(x_bl)], [x_br, get_chord_y_bot(x_br)],
                        [x_tr, get_chord_y_top(x_tr)], [x_tl, get_chord_y_top(x_tl)]
                    ]
                else:
                    x_br, x_bl = px_bot + w_half, px_bot - w_half
                    x_tl, x_tr = px_top - w_half, px_top + w_half
                    pts = [
                        [x_br, get_chord_y_bot(x_br)], [x_bl, get_chord_y_bot(x_bl)],
                        [x_tl, get_chord_y_top(x_tl)], [x_tr, get_chord_y_top(x_tr)]
                    ]
                    
                poly = plt.Polygon(pts, facecolor='#f1c40f', edgecolor='black', linewidth=1.2, zorder=3)
                ax.add_patch(poly)
                
                dx_line, dy_line = px_top - px_bot, py_top - py_bot
                diag_ang = math.degrees(math.atan2(dy_line, dx_line))
                
                t_slope = get_slope(get_y_top, px_top)
                b_slope = get_slope(get_y_bot, px_bot)
                
                t_intersect = abs(diag_ang - t_slope) % 180
                if t_intersect > 90: t_intersect = 180 - t_intersect
                d_top_angle = int(round(abs(90.0 - t_intersect)))
                
                b_intersect = abs(diag_ang - b_slope) % 180
                if b_intersect > 90: b_intersect = 180 - b_intersect
                d_bot_angle = int(round(abs(90.0 - b_intersect)))

                mx, my = (px_bot + px_top) / 2, (py_bot + py_top) / 2
                draw_dim_text(ax, mx, my, f"L:{diag_l:.1f} ({d_top_angle}°/{d_bot_angle}°)", angle=diag_ang, color='#8B0000', fontsize=11)

                raw_data.append({
                    "구분": "살대", "품명": f"{d_od}mm 파이프",
                    "재단기장(L)": round(diag_l, 1), "상단 가공각(°)": d_top_angle, "하단 가공각(°)": d_bot_angle
                })

            if is_double_bot:
                def draw_custom_diag(left_x, right_x, gubun_name, is_forward):
                    dx = right_x - left_x
                    if dx <= 0: return
                    
                    if gubun_name == "상단살대":
                        y_bot_est = H_mid_top
                        y_top_est = get_chord_y_top((left_x+right_x)/2)
                    else:
                        y_top_est = H_mid_bot
                        y_bot_est = get_chord_y_bot((left_x+right_x)/2)
                        
                    v_len = abs(y_top_est - y_bot_est)
                    diag_len = math.hypot(dx, v_len)
                    sin_theta = v_len / diag_len if diag_len > 0 else 1
                    w_half = (d_od / 2) / sin_theta if sin_theta > 0.01 else d_od / 2
                    
                    px_bot = left_x + w_half if is_forward else right_x - w_half
                    px_top = right_x - w_half if is_forward else left_x + w_half
                    
                    if gubun_name == "상단살대":
                        py_bot = H_mid_top
                        py_top = get_chord_y_top(px_top)
                    else:
                        py_top = H_mid_bot
                        py_bot = get_chord_y_bot(px_bot)
                        
                    if py_top <= py_bot: return 
                    
                    diag_l = math.hypot(px_top - px_bot, py_top - py_bot)
                    
                    if is_forward:
                        x_bl, x_br = px_bot - w_half, px_bot + w_half
                        x_tr, x_tl = px_top + w_half, px_top - w_half
                        if gubun_name == "상단살대":
                            pts = [
                                [x_bl, H_mid_top], [x_br, H_mid_top],
                                [x_tr, get_chord_y_top(x_tr)], [x_tl, get_chord_y_top(x_tl)]
                            ]
                        else:
                            pts = [
                                [x_bl, get_chord_y_bot(x_bl)], [x_br, get_chord_y_bot(x_br)],
                                [x_tr, H_mid_bot], [x_tl, H_mid_bot]
                            ]
                    else:
                        x_br, x_bl = px_bot + w_half, px_bot - w_half
                        x_tl, x_tr = px_top - w_half, px_top + w_half
                        if gubun_name == "상단살대":
                            pts = [
                                [x_br, H_mid_top], [x_bl, H_mid_top],
                                [x_tl, get_chord_y_top(x_tl)], [x_tr, get_chord_y_top(x_tr)]
                            ]
                        else:
                            pts = [
                                [x_br, get_chord_y_bot(x_br)], [x_bl, get_chord_y_bot(x_bl)],
                                [x_tl, H_mid_bot], [x_tr, H_mid_bot]
                            ]
                            
                    poly = plt.Polygon(pts, facecolor='#f1c40f', edgecolor='black', linewidth=1.2, zorder=3)
                    ax.add_patch(poly)
                    
                    dx_line, dy_line = px_top - px_bot, py_top - py_bot
                    diag_ang = math.degrees(math.atan2(dy_line, dx_line))
                    
                    if gubun_name == "상단살대":
                        t_slope = get_slope(get_y_top, px_top)
                        b_slope = 0.0
                    else:
                        t_slope = 0.0 
                        b_slope = get_slope(get_y_bot, px_bot)
                        
                    t_intersect = abs(diag_ang - t_slope) % 180
                    if t_intersect > 90: t_intersect = 180 - t_intersect
                    d_top_angle = int(round(abs(90.0 - t_intersect)))
                    
                    b_intersect = abs(diag_ang - b_slope) % 180
                    if b_intersect > 90: b_intersect = 180 - b_intersect
                    d_bot_angle = int(round(abs(90.0 - b_intersect)))
                    
                    mx, my = (px_bot + px_top)/2, (py_bot + py_top)/2
                    draw_dim_text(ax, mx, my, f"L:{diag_l:.1f} ({d_top_angle}°/{d_bot_angle}°)", angle=diag_ang, color='#8B0000', fontsize=10)
                    
                    raw_data.append({
                        "구분": gubun_name, "품명": f"{d_od}mm 파이프",
                        "재단기장(L)": round(diag_l, 1), "상단 가공각(°)": d_top_angle, "하단 가공각(°)": d_bot_angle
                    })

                is_forward_u = not (is_half or i < mid_idx)
                draw_custom_diag(wx_start, wx_end, "상단살대", is_forward_u)

                if is_half: is_forward_l = (i % 2 == 0)
                else:
                    if i < mid_idx: is_forward_l = ((mid_idx - 1 - i) % 2 != 0) 
                    else: is_forward_l = ((i - mid_idx) % 2 == 0)
                draw_custom_diag(wx_start, wx_end, "하단살대", is_forward_l)

            else:
                if is_half or i < mid_idx: draw_diag(wx_start, wx_end, False)
                else: draw_diag(wx_start, wx_end, True)

    if has_tie and H_tie > 0:
        y_tie_top = H_tie + m_od / 2
        
        low, high = 0.0, S/2
        for _ in range(50):
            mid = (low + high) / 2
            y_inner = get_y_bot(mid)
            if y_inner < y_tie_top: low = mid
            else: high = mid
        x_left_in = (low + high) / 2
        
        if is_half:
            x_right_in = S - m_od / 2
        else:
            low, high = S/2, S
            for _ in range(50):
                mid = (low + high) / 2
                y_inner = get_y_bot(mid)
                if y_inner > y_tie_top:  
                    low = mid
                else:
                    high = mid
            x_right_in = (low + high) / 2

        tie_length_inner = x_right_in - x_left_in
        
        if tie_length_inner > 0:
            ax.add_patch(patches.Rectangle((x_left_in, H_tie - m_od/2), tie_length_inner, m_od, facecolor='#8e44ad', edgecolor='black', zorder=6))
            ax.annotate(f"수평재 내경 L:{tie_length_inner:.1f}", xy=((x_left_in+x_right_in)/2, H_tie), xytext=((x_left_in+x_right_in)/2, H_tie - 120),
                        arrowprops=dict(arrowstyle='<->', color='purple', lw=3), ha='center', fontsize=16, fontweight='bold', color='purple')
            
            raw_data.append({
                "구분": "수평재", "품명": f"{m_od}mm 파이프", "재단기장(L)": round(tie_length_inner, 1),
                "상단 가공각(°)": 0, "하단 가공각(°)": 0
            })

        valid_inner_xs = []
        for i in range(1, divs):
            if i == mid_idx and not is_half:
                if type_choice != "5": continue
            x = v_centers_x[i]
            if get_y_bot(x) > (H_tie + m_od/2) + 10:
                valid_inner_xs.append((i, x))
                
        for i, x in valid_inner_xs:
            is_r_curr = (i == mid_idx) and not is_half
            curr_v_od = r_od if is_r_curr else v_od
            
            x_l, x_r = x - curr_v_od / 2, x + curr_v_od / 2
            
            yt_l, yt_r = get_y_bot(x_l), get_y_bot(x_r)
            
            if x_l < S/2 < x_r and not is_arch:
                poly_top_inner = [[x_r, yt_r], [S/2, get_y_bot(S/2)], [x_l, yt_l]]
            else:
                poly_top_inner = [[x_r, yt_r], [x_l, yt_l]]
            
            y_bot = H_tie + m_od/2
            v_len = max(yt_l, yt_r) - y_bot
            
            pts_inner = [[x_l, y_bot], [x_r, y_bot]] + poly_top_inner
            ax.add_patch(patches.Polygon(pts_inner, facecolor='#2ecc71', edgecolor='black', zorder=4))
            t_angle = int(round(abs(get_slope(get_y_bot, x))))
            
            stagger_bot = 500 if i % 2 == 0 else 800
            if is_r_curr: stagger_bot = 1100 
            my_bot = H_tie - stagger_bot
            
            ax.plot([x, x], [H_tie - m_od/2, my_bot + 180], color='darkgreen', linestyle=':', lw=1.5, zorder=1)
            draw_dim_text(ax, x, my_bot, f"L:{v_len:.1f} (상:{t_angle}°/하:0°)", angle=90, color='darkgreen')
            
            raw_data.append({
                "구분": "수평내부다대", "품명": f"{curr_v_od}mm 파이프", "재단기장(L)": round(v_len, 1),
                "상단 가공각(°)": t_angle, "하단 가공각(°)": 0
            })
            
        if type_choice == "5":
            tie_bot_y = H_tie - m_od/2
            dim_h_x = S/2 - r_od/2 - 250
            if is_half: dim_h_x = S/2 - 250
            
            ax.plot([dim_h_x - 100, dim_h_x + 100], [0, 0], color='black', lw=1.5, zorder=1)
            ax.annotate("", xy=(dim_h_x, tie_bot_y), xytext=(dim_h_x, 0),
                        arrowprops=dict(arrowstyle='<->', color='#d35400', lw=2.5))
            draw_dim_text(ax, dim_h_x, tie_bot_y / 2, f"수평보 하단 높이: {tie_bot_y:.1f}", angle=90, color='#d35400', fontsize=13)

        valid_inner_x_vals = [x for _, x in valid_inner_xs]
        base_intervals = [x_left_in] + valid_inner_x_vals + [x_right_in]
        if not is_half: base_intervals.append(S/2)
        inner_intervals = sorted(list(set(base_intervals)))
        
        for i in range(len(inner_intervals)-1):
            if i == 0:
                continue 
            if not is_half and i == len(inner_intervals) - 2:
                continue 

            lx, rx = inner_intervals[i], inner_intervals[i+1]
            
            l_v_od = r_od if (round(lx, 1) == round(S/2, 1) and not is_half) else v_od
            r_v_od = r_od if (round(rx, 1) == round(S/2, 1) and not is_half) else v_od
            wx_start, wx_end = lx + l_v_od/2 + offset_mm, rx - r_v_od/2 - offset_mm
            y_bot_limit = H_tie + m_od/2
            
            if wx_end > wx_start + 10: 
                is_forward_tie = not (is_half or lx < S/2)
                
                dy_est = get_y_bot(wx_start) - y_bot_limit if not is_forward_tie else get_y_bot(wx_end) - y_bot_limit
                dx_est = wx_end - wx_start
                diag_len_est = math.hypot(dx_est, dy_est)
                sin_theta = abs(dy_est) / diag_len_est if diag_len_est > 0 else 1
                w_half_tie = (d_od / 2) / sin_theta if sin_theta > 0.01 else d_od / 2
                
                px_bot = wx_start + w_half_tie if is_forward_tie else wx_end - w_half_tie
                px_top = wx_end - w_half_tie if is_forward_tie else wx_start + w_half_tie
                
                py_bot = y_bot_limit
                py_top = get_y_bot(px_top) 
                
                diag_l = math.hypot(px_top - px_bot, py_top - py_bot)
                
                if is_forward_tie:
                    x_bl, x_br = wx_start, wx_start + 2*w_half_tie
                    x_tr, x_tl = wx_end, wx_end - 2*w_half_tie
                    pts = [
                        [x_bl, y_bot_limit], [x_br, y_bot_limit],
                        [x_tr, get_y_bot(x_tr)], [x_tl, get_y_bot(x_tl)]
                    ]
                else:
                    x_br, x_bl = wx_end, wx_end - 2*w_half_tie
                    x_tl, x_tr = wx_start, wx_start + 2*w_half_tie
                    pts = [
                        [x_br, y_bot_limit], [x_bl, y_bot_limit],
                        [x_tl, get_y_bot(x_tl)], [x_tr, get_y_bot(x_tr)]
                    ]
                    
                poly = plt.Polygon(pts, facecolor='#f1c40f', edgecolor='black', linewidth=1.2, zorder=3)
                ax.add_patch(poly)
                
                dx_diff, dy_diff = px_top - px_bot, py_top - py_bot
                diag_ang = math.degrees(math.atan2(dy_diff, dx_diff))
                
                t_slope = get_slope(get_y_bot, px_top)
                b_slope = 0.0
                
                t_intersect = abs(diag_ang - t_slope) % 180
                if t_intersect > 90: t_intersect = 180 - t_intersect
                d_top_angle = int(round(abs(90.0 - t_intersect)))
                
                b_intersect = abs(diag_ang - b_slope) % 180
                if b_intersect > 90: b_intersect = 180 - b_intersect
                d_bot_angle = int(round(abs(90.0 - b_intersect)))
                
                mx, my = (px_bot + px_top)/2, (py_bot + py_top)/2
                draw_dim_text(ax, mx, my, f"L:{diag_l:.1f} ({d_top_angle}°/{d_bot_angle}°)", angle=diag_ang, color='#b8860b', fontsize=11)
                
                raw_data.append({
                    "구분": "수평내부살대", "품명": f"{d_od}mm 파이프", "재단기장(L)": round(diag_l, 1),
                    "상단 가공각(°)": d_top_angle, "하단 가공각(°)": d_bot_angle
                })

    dim_y = -1500 if is_double_bot else -350
    ax.plot([0, S], [dim_y, dim_y], color='black', lw=2, zorder=10)
    ax.plot([0, 0], [dim_y - 25, dim_y + 25], color='black', lw=2, zorder=10)
    ax.plot([S, S], [dim_y - 25, dim_y + 25], color='black', lw=2, zorder=10)

    ticks_x = [0] + [v_centers_x[i] for i in range(1, divs)] + [S]
    
    for i in range(divs):
        tx1, tx2 = ticks_x[i], ticks_x[i+1]
        cx = (tx1 + tx2) / 2
        interval_len = tx2 - tx1
        
        if i > 0: 
            ax.plot([tx1, tx1], [dim_y - 20, dim_y + 20], color='black', lw=1.5, zorder=10)
            ax.plot([tx1, tx1], [0, dim_y], color='gray', linestyle=':', lw=1.5, zorder=1)
        
        f_size = 12 if interval_len > 300 else 10
        ax.text(cx, dim_y + 40, f"{interval_len:.1f}", ha='center', va='center', fontsize=f_size, color='navy', fontweight='bold')

    for tx in ticks_x:
        ax.text(tx, dim_y - 45, f"{tx:.1f}", ha='center', va='top', fontsize=11, color='#d35400', fontweight='bold')

    ax.plot([S, S], [0, dim_y], color='gray', linestyle=':', lw=1.5, zorder=1)
    ax.text(S/2, dim_y - 120, f"전체 스판 : {S:.1f} mm", ha='center', va='center', fontsize=18, fontweight='bold', color='black')

    ax.set_xlim(-200, S + 200)
    ax.set_ylim(dim_y - 300, H_cen + 1200) 
    ax.set_aspect('equal')
    ax.axis('off') 
    
    info_text = f"스판: {span_cm}cm | 등분: {divs} (자간: {interval_len/10:.1f}cm)"
    if has_tie: info_text += f" | 수평재 높이: {h_tie_cm}cm"
    if is_double_bot: info_text += f" | 밑더블 외경 높이: {h_outer_cm}cm"
    
    plt.title(f"트러스 도면 ({t_name})\n{info_text}", fontsize=24, fontweight='bold', pad=20)
    
    # Save to memory
    file_prefix = f"Truss_{t_name.replace(' ', '')}_{int(span_cm)}"
    
    pdf_buffer = io.BytesIO()
    plt.savefig(pdf_buffer, format='pdf', bbox_inches='tight')
    pdf_buffer.seek(0)
    
    excel_path = f"{file_prefix}_재단표.xlsx"
    save_formatted_excel(raw_data, excel_path)
    
    with open(excel_path, "rb") as f:
        excel_bytes = f.read()
    
    return fig, pdf_buffer, excel_bytes, f"{file_prefix}.pdf", excel_path


# ==============================================================================
# [2] 벽사다리/용마루 시스템 관련 함수
# ==============================================================================
def get_6m_count(total_cm):
    return math.ceil(total_cm / 600.0)

def set_excel_style(ws):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
    fill_main = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")   
    fill_v = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")      
    fill_angle = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  
    fill_sep = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")    
    
    header_font = Font(color="FFFFFF", bold=True)
    cut_size_font = Font(color="FF0000", bold=True) 
    angle_text_font = Font(color="002060", bold=True) 

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = (max_length * 1.6) + 6

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        is_header = (row_idx == 0)
        item_name = str(row[0].value) if row[0].value else ""
        
        for col_idx, cell in enumerate(row):
            cell.border = thin_border
            cell.alignment = center_align
            
            if is_header:
                cell.fill = header_fill
                cell.font = header_font
            else:
                if "---" in item_name:
                    cell.fill = fill_sep
                elif "다대" in item_name or "수직" in item_name:
                    cell.fill = fill_v
                elif "살대" in item_name or "사재" in item_name or "°" in item_name or "가공각" in item_name:
                    cell.fill = fill_angle
                    if col_idx == 0:
                        cell.font = angle_text_font
                elif "상하현재" in item_name or "스나기" in item_name or "안내" in item_name:
                    cell.fill = fill_main
                
                if col_idx == 3 and "---" not in item_name:
                    cell.font = cut_size_font

def draw_pipe(ax, x1, y1, x2, y2, t, zorder=1, facecolor='white'):
    dx = x2 - x1
    dy = y2 - y1
    length = math.hypot(dx, dy)
    if length == 0: return
    
    nx = -dy / length * (t / 2)
    ny = dx / length * (t / 2)
    
    poly = plt.Polygon(
        [[x1+nx, y1+ny], [x2+nx, y2+ny], [x2-nx, y2-ny], [x1-nx, y1-ny]],
        facecolor=facecolor, edgecolor='black', linewidth=1.2, zorder=zorder
    )
    ax.add_patch(poly)

def draw_diag_poly(ax, x_start_tip, x_end_tip, y_bot, y_top, w_half, is_forward, zorder=1, facecolor='yellow'):
    t_h = 2 * w_half  
    if is_forward: 
        pts = [
            [x_start_tip, y_bot],             
            [x_start_tip + t_h, y_bot],       
            [x_end_tip, y_top],               
            [x_end_tip - t_h, y_top]          
        ]
    else:          
        pts = [
            [x_start_tip, y_top],             
            [x_start_tip + t_h, y_top],       
            [x_end_tip, y_bot],               
            [x_end_tip - t_h, y_bot]          
        ]
    poly = plt.Polygon(pts, facecolor=facecolor, edgecolor='black', linewidth=1.2, zorder=zorder)
    ax.add_patch(poly)

def run_ladder_system(params):
    """벽사다리 통합 산출 시스템 (Streamlit 파라미터 매핑)"""
    L_cm = params['L_cm']
    W_cm = params['W_cm']
    H_truss_cm = params['H_truss_cm']
    total_sets_sub = params['total_sets_sub']
    total_sets_main = params['total_sets_main']
    offset_mm = params['offset_mm']
    offset_cm = offset_mm / 10.0

    H_ridge_cm = params['H_ridge_cm']
    ridge_deduct_mm = params['ridge_deduct_mm']
    ridge_deduct_cm = ridge_deduct_mm / 10.0
    total_sets_ridge = params['total_sets_ridge']

    wall_snagi_mm = params['wall_snagi_mm']
    wall_deduct_cm = wall_snagi_mm / 10.0
    wall_snagi_cm = wall_snagi_mm / 10.0

    p_sub_main = params['p_sub_main']
    p_sub_sub = params['p_sub_sub']
    p_main_main = params['p_main_main']
    p_main_snagi = params['p_main_snagi']
    p_main_v = params['p_main_v']
    p_main_diag = params['p_main_diag']
    p_ridge_main = params['p_ridge_main']
    p_ridge_v = params['p_ridge_v']
    p_ridge_diag = params['p_ridge_diag']
    
    max_span_main = 380.0 

    t_sub_main_cm = p_sub_main / 10.0
    t_sub_sub_cm = p_sub_sub / 10.0
    t_main_main_cm = p_main_main / 10.0
    t_main_snagi_cm = p_main_snagi / 10.0
    t_main_v_cm = p_main_v / 10.0
    t_main_diag_cm = p_main_diag / 10.0
    
    t_ridge_main_cm = p_ridge_main / 10.0
    t_ridge_v_cm = p_ridge_v / 10.0
    t_ridge_diag_cm = p_ridge_diag / 10.0

    actual_sub_v_len = W_cm - (2 * t_sub_main_cm)
    actual_main_v_len = H_truss_cm - (2 * t_main_main_cm)
    actual_ridge_v_len = H_ridge_cm - (2 * t_ridge_main_cm)

    n_sec_m = math.ceil(L_cm / max_span_main)
    gap_m = L_cm / n_sec_m
    sub_div = 4
    sub_gap = gap_m / sub_div
    dx_m = offset_cm
    
    n_sec_s = n_sec_m * sub_div 
    gap_s = (L_cm - t_sub_sub_cm) / n_sec_s
    dx_s = offset_cm 
    
    def calc_diag(spacing, v_len, left_r, right_r, t_diag):
        eff_spacing = spacing - left_r - right_r - (2 * offset_cm)
        dx_center = eff_spacing
        W_half = 0
        for _ in range(10): 
            diag_len = math.hypot(dx_center, v_len)
            sin_theta = v_len / diag_len if diag_len > 0 else 1
            W_half = (t_diag / 2) / sin_theta
            dx_center = eff_spacing - 2 * W_half
            if dx_center < 0: dx_center = 0.1
            
        actual_diag = math.sqrt(dx_center**2 + v_len**2)
        angle_deg = math.degrees(math.atan2(v_len, dx_center))
        cut_angle = 90.0 - angle_deg
        return actual_diag, cut_angle, angle_deg, W_half

    actual_sub_diag_len, sub_cut_angle, angle_s_deg, w_half_s = calc_diag(gap_s, actual_sub_v_len, t_sub_sub_cm/2, t_sub_sub_cm/2, t_sub_sub_cm)

    diag_len_snagi, cut_angle_snagi, angle_m_snagi_deg, w_half_m_snagi = calc_diag(sub_gap, actual_main_v_len, t_main_snagi_cm/2, t_main_v_cm/2, t_main_diag_cm)
    diag_len_norm, cut_angle_norm, angle_m_norm_deg, w_half_m_norm = calc_diag(sub_gap, actual_main_v_len, t_main_v_cm/2, t_main_v_cm/2, t_main_diag_cm)

    L_ridge_mid = gap_m - ridge_deduct_cm
    L_ridge_end = gap_m - 1.5 * ridge_deduct_cm
    L_ridge_single = gap_m - 2.0 * ridge_deduct_cm

    ridge_single_count = 1 if n_sec_m == 1 else 0
    ridge_end_count = 2 if n_sec_m >= 2 else 0
    ridge_mid_count = max(0, n_sec_m - 2)
    
    S_norm = sub_gap
    S_mid_deep = sub_gap - (ridge_deduct_cm / 2)
    S_end_deep = sub_gap - ridge_deduct_cm

    diag_norm, cut_norm, _, w_half_r_norm = calc_diag(S_norm, actual_ridge_v_len, t_ridge_v_cm/2, t_ridge_v_cm/2, t_ridge_diag_cm)
    diag_mid_c, cut_mid_c, _, w_half_r_mid = calc_diag(S_mid_deep, actual_ridge_v_len, ridge_deduct_cm/2, t_ridge_v_cm/2, t_ridge_diag_cm)
    diag_end_c, cut_end_c, _, w_half_r_end = calc_diag(S_end_deep, actual_ridge_v_len, ridge_deduct_cm/2, ridge_deduct_cm/2, t_ridge_diag_cm)

    L_wall_single = gap_m - 2.0 * wall_snagi_cm
    L_wall_end = gap_m - 1.5 * wall_snagi_cm
    L_wall_mid = gap_m - wall_snagi_cm
    wall_mid_count = max(0, n_sec_m - 2)

    data = [
        ["[보강사다리] 상하현재", f"{p_sub_main}mm", 2 * total_sets_sub, L_cm, round((L_cm*2)*total_sets_sub, 1), get_6m_count((L_cm*2)*total_sets_sub)],
        ["[보강사다리] 수직재(다대)", f"{p_sub_sub}mm", (n_sec_s+1) * total_sets_sub, round(actual_sub_v_len, 1), round(((n_sec_s+1)*actual_sub_v_len)*total_sets_sub, 1), get_6m_count(((n_sec_s+1)*actual_sub_v_len)*total_sets_sub)],
        [f"[보강사다리] 사재(가공각:{int(round(sub_cut_angle, 0))}°)", f"{p_sub_sub}mm", n_sec_s * total_sets_sub, round(actual_sub_diag_len,1), round((actual_sub_diag_len*n_sec_s)*total_sets_sub, 1), get_6m_count((actual_sub_diag_len*n_sec_s)*total_sets_sub)],
        ["------------------", "----------", "---", "---", "---", "---"],
        
        ["[메인사다리] 상하현재", f"{p_main_main}mm", 2 * total_sets_main, L_cm, round((L_cm*2)*total_sets_main, 1), get_6m_count((L_cm*2)*total_sets_main)],
        ["[메인사다리] 스나기", f"{p_main_snagi}mm", (n_sec_m+1) * total_sets_main, H_truss_cm+30, round(((n_sec_m+1)*(H_truss_cm+30))*total_sets_main, 1), get_6m_count(((n_sec_m+1)*(H_truss_cm+30))*total_sets_main)],
        ["[메인사다리] 수직다대", f"{p_main_v}mm", (n_sec_m*sub_div+1) * total_sets_main, round(actual_main_v_len, 1), round(((n_sec_m*sub_div+1)*actual_main_v_len)*total_sets_main, 1), get_6m_count(((n_sec_m*sub_div+1)*actual_main_v_len)*total_sets_main)],
        
        [f"[메인사다리] 살대-스나기접합({int(round(cut_angle_snagi, 0))}°)", f"{p_main_diag}mm", 2 * n_sec_m * total_sets_main, round(diag_len_snagi,1), round((diag_len_snagi*2*n_sec_m)*total_sets_main, 1), get_6m_count((diag_len_snagi*2*n_sec_m)*total_sets_main)]
    ]
    
    if sub_div > 2:
        qty_norm_diag = (sub_div - 2) * n_sec_m * total_sets_main
        data.append(
            [f"[메인사다리] 살대-일반({int(round(cut_angle_norm, 0))}°)", f"{p_main_diag}mm", qty_norm_diag, round(diag_len_norm,1), round((diag_len_norm*qty_norm_diag), 1), get_6m_count(diag_len_norm*qty_norm_diag)]
        )
    data.append(["------------------", "----------", "---", "---", "---", "---"])

    if ridge_single_count > 0:
        data.extend([
            ["[용마루] 단일세트 상하현재", f"{p_ridge_main}mm", 2 * total_sets_ridge, round(L_ridge_single,1), round((L_ridge_single*2)*total_sets_ridge, 1), get_6m_count((L_ridge_single*2)*total_sets_ridge)],
            ["[용마루] 단일세트 다대", f"{p_ridge_v}mm", 5 * total_sets_ridge, round(actual_ridge_v_len,1), round((actual_ridge_v_len*5)*total_sets_ridge, 1), get_6m_count((actual_ridge_v_len*5)*total_sets_ridge)],
            [f"[용마루] 단일세트 살대-일반({int(round(cut_norm, 0))}°)", f"{p_ridge_diag}mm", 2 * total_sets_ridge, round(diag_norm,1), round((diag_norm*2)*total_sets_ridge, 1), get_6m_count((diag_norm*2)*total_sets_ridge)],
            [f"[용마루] 단일세트 살대-양끝공제({int(round(cut_end_c, 0))}°)", f"{p_ridge_diag}mm", 2 * total_sets_ridge, round(diag_end_c,1), round((diag_end_c*2)*total_sets_ridge, 1), get_6m_count((diag_end_c*2)*total_sets_ridge)],
        ])
    else:
        if ridge_end_count > 0:
            data.extend([
                ["[용마루] 양끝세트 상하현재", f"{p_ridge_main}mm", 2 * ridge_end_count * total_sets_ridge, round(L_ridge_end,1), round((L_ridge_end*2*ridge_end_count)*total_sets_ridge, 1), get_6m_count((L_ridge_end*2*ridge_end_count)*total_sets_ridge)],
                ["[용마루] 양끝세트 다대", f"{p_ridge_v}mm", 5 * ridge_end_count * total_sets_ridge, round(actual_ridge_v_len,1), round((actual_ridge_v_len*5*ridge_end_count)*total_sets_ridge, 1), get_6m_count((actual_ridge_v_len*5*ridge_end_count)*total_sets_ridge)],
                [f"[용마루] 양끝세트 살대-일반({int(round(cut_norm, 0))}°)", f"{p_ridge_diag}mm", 2 * ridge_end_count * total_sets_ridge, round(diag_norm,1), round((diag_norm*2*ridge_end_count)*total_sets_ridge, 1), get_6m_count((diag_norm*2*ridge_end_count)*total_sets_ridge)],
                [f"[용마루] 양끝세트 살대-안쪽공제({int(round(cut_mid_c, 0))}°)", f"{p_ridge_diag}mm", 1 * ridge_end_count * total_sets_ridge, round(diag_mid_c,1), round((diag_mid_c*1*ridge_end_count)*total_sets_ridge, 1), get_6m_count((diag_mid_c*1*ridge_end_count)*total_sets_ridge)],
                [f"[용마루] 양끝세트 살대-끝쪽공제({int(round(cut_end_c, 0))}°)", f"{p_ridge_diag}mm", 1 * ridge_end_count * total_sets_ridge, round(diag_end_c,1), round((diag_end_c*1*ridge_end_count)*total_sets_ridge, 1), get_6m_count((diag_end_c*1*ridge_end_count)*total_sets_ridge)],
            ])
        if ridge_mid_count > 0:
            data.extend([
                ["[용마루] 중간세트 상하현재", f"{p_ridge_main}mm", 2 * ridge_mid_count * total_sets_ridge, round(L_ridge_mid,1), round((L_ridge_mid*2*ridge_mid_count)*total_sets_ridge, 1), get_6m_count((L_ridge_mid*2*ridge_mid_count)*total_sets_ridge)],
                ["[용마루] 중간세트 다대", f"{p_ridge_v}mm", 5 * ridge_mid_count * total_sets_ridge, round(actual_ridge_v_len,1), round((actual_ridge_v_len*5*ridge_mid_count)*total_sets_ridge, 1), get_6m_count((actual_ridge_v_len*5*ridge_mid_count)*total_sets_ridge)],
                [f"[용마루] 중간세트 살대-일반({int(round(cut_norm, 0))}°)", f"{p_ridge_diag}mm", 2 * ridge_mid_count * total_sets_ridge, round(diag_norm,1), round((diag_norm*2*ridge_mid_count)*total_sets_ridge, 1), get_6m_count((diag_norm*2*ridge_mid_count)*total_sets_ridge)],
                [f"[용마루] 중간세트 살대-공제부({int(round(cut_mid_c, 0))}°)", f"{p_ridge_diag}mm", 2 * ridge_mid_count * total_sets_ridge, round(diag_mid_c,1), round((diag_mid_c*2*ridge_mid_count)*total_sets_ridge, 1), get_6m_count((diag_mid_c*2*ridge_mid_count)*total_sets_ridge)],
            ])
    
    data.append(["------------------", "----------", "---", "---", "---", "---"])
    if n_sec_m == 1:
        data.append(["[안내] 벽사다리 단일 가로절단", "-", "-", round(L_wall_single, 1), "-", f"스나기({wall_snagi_mm}mm) 기준 적용"])
    else:
        data.append(["[안내] 벽사다리 양끝 가로절단", "-", "-", round(L_wall_end, 1), "-", f"스나기({wall_snagi_mm}mm) 기준 적용"])
        if wall_mid_count > 0:
            data.append(["[안내] 벽사다리 중간 가로절단", "-", "-", round(L_wall_mid, 1), "-", f"스나기({wall_snagi_mm}mm) 기준 적용"])
    
    df = pd.DataFrame(data, columns=["항목", "규격", "수량(개/줄)", "단위길이(cm)", "총연장(cm)", "6m본수/비고"])
    excel_file = f"종합산출표_{int(L_cm)}cm.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='전체자재표')
        set_excel_style(writer.sheets['전체자재표'])

    pdf_file_name = f"종합상세도면_{int(L_cm)}cm.pdf"
    pdf_buffer = io.BytesIO()
    with PdfPages(pdf_buffer) as pdf:
        fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(max(26, L_cm/30), 26), gridspec_kw={'height_ratios': [1, 1.2, 1]})
        plt.subplots_adjust(hspace=0.25, top=0.95, bottom=0.05)

        y_top_center_s = W_cm - (t_sub_main_cm / 2)
        y_bot_center_s = 0 + (t_sub_main_cm / 2)
        y_inner_top_s = W_cm - t_sub_main_cm
        y_inner_bot_s = t_sub_main_cm

        ax1.annotate('', xy=(0, W_cm*1.3), xytext=(L_cm, W_cm*1.3), arrowprops=dict(arrowstyle='<->', color='red', lw=2))
        ax1.text(L_cm/2, W_cm*1.35, f"총기장: {L_cm}cm", ha='center', color='red', weight='bold', fontsize=28) 

        draw_pipe(ax1, 0, y_top_center_s, L_cm, y_top_center_s, t_sub_main_cm, zorder=3, facecolor='#C0C0C0')
        draw_pipe(ax1, 0, y_bot_center_s, L_cm, y_bot_center_s, t_sub_main_cm, zorder=3, facecolor='#C0C0C0')
        
        v_centers_s = [(t_sub_sub_cm / 2) + i * gap_s for i in range(n_sec_s + 1)]
        
        for i in range(n_sec_s + 1):
            x = v_centers_s[i]
            draw_pipe(ax1, x, y_inner_bot_s, x, y_inner_top_s, t_sub_sub_cm, zorder=2, facecolor='#000080')
            
            if i < n_sec_s:
                nx_val = v_centers_s[i+1]
                ax1.text((x + nx_val)/2, -W_cm*0.2, f"{round(gap_s,1)}cm", ha='center', fontsize=20, color='black') 
                
                start_tip_s = x + (t_sub_sub_cm / 2) + dx_s
                end_tip_s = nx_val - (t_sub_sub_cm / 2) - dx_s
                
                if i % 2 == 0: 
                    draw_diag_poly(ax1, start_tip_s, end_tip_s, y_inner_bot_s, y_inner_top_s, w_half_s, True, zorder=1, facecolor='yellow')
                    ax1.text((start_tip_s + end_tip_s)/2, W_cm/2, f"{round(actual_sub_diag_len,1)}cm\n({int(round(sub_cut_angle, 0))}°)", ha='center', va='center', fontsize=18, color='blue', rotation=angle_s_deg) 
                else:          
                    draw_diag_poly(ax1, start_tip_s, end_tip_s, y_inner_bot_s, y_inner_top_s, w_half_s, False, zorder=1, facecolor='yellow')
                    ax1.text((start_tip_s + end_tip_s)/2, W_cm/2, f"{round(actual_sub_diag_len,1)}cm\n({int(round(sub_cut_angle, 0))}°)", ha='center', va='center', fontsize=18, color='blue', rotation=-angle_s_deg) 

        ax1.text(-gap_s*0.1, W_cm/2, f"다대 실절단\n{round(actual_sub_v_len,1)}cm", va='center', ha='right', color='black', fontsize=24) 
        ax1.set_title(f"1. 보강사다리 상세 (폭: {W_cm}cm)", fontsize=36, fontweight='bold', pad=30) 
        ax1.set_xlim(-L_cm*0.08, L_cm*1.08)
        ax1.set_ylim(-W_cm*0.4, W_cm*1.6)
        ax1.axis('off'); ax1.set_aspect('equal')

        y_top_center_m = H_truss_cm - (t_main_main_cm / 2)
        y_bot_center_m = 0 + (t_main_main_cm / 2)
        y_inner_top_m = H_truss_cm - t_main_main_cm
        y_inner_bot_m = t_main_main_cm

        y_total_line = H_truss_cm + 65
        ax2.annotate('', xy=(0, y_total_line), xytext=(L_cm, y_total_line), arrowprops=dict(arrowstyle='<->', color='red', lw=2))
        ax2.text(L_cm/2, y_total_line + 5, f"총기장: {L_cm}cm", ha='center', color='red', weight='bold', fontsize=28) 

        draw_pipe(ax2, 0, y_top_center_m, L_cm, y_top_center_m, t_main_main_cm, zorder=3, facecolor='#C0C0C0')
        draw_pipe(ax2, 0, y_bot_center_m, L_cm, y_bot_center_m, t_main_main_cm, zorder=3, facecolor='#C0C0C0')
        
        for i in range(n_sec_m + 1):
            cx = i * gap_m
            draw_pipe(ax2, cx, -30, cx, H_truss_cm, t_main_snagi_cm, zorder=5, facecolor='purple')
            
            if i == 0:
                ax2.annotate('', xy=(cx - t_main_snagi_cm, 0), xytext=(cx - t_main_snagi_cm, -30), arrowprops=dict(arrowstyle='<->', color='purple', lw=1.5))
                ax2.text(cx - t_main_snagi_cm - 2, -15, "30cm", ha='right', va='center', color='purple', fontsize=20, weight='bold') 

            if i < n_sec_m:
                ax2.text(cx + gap_m/2, H_truss_cm + 25, f"구분간격(벽사다리 스팬): {round(gap_m,1)}cm", ha='center', color='black', weight='bold', fontsize=24) 
                for j in range(1, sub_div + 1):
                    px, curx = cx + ((j-1) * sub_gap), cx + (j * sub_gap)
                    
                    if j < sub_div:
                        draw_pipe(ax2, curx, y_inner_bot_m, curx, y_inner_top_m, t_main_v_cm, zorder=4, facecolor='#000080')
                    
                    ax2.text((px + curx)/2, -20, f"자간:\n{round(sub_gap,1)}cm", ha='center', fontsize=18, color='gray') 
                    
                    r_left = (t_main_snagi_cm / 2) if j == 1 else (t_main_v_cm / 2)
                    r_right = (t_main_snagi_cm / 2) if j == sub_div else (t_main_v_cm / 2)
                    
                    is_snagi_adj = (j == 1 or j == sub_div)
                    w_half = w_half_m_snagi if is_snagi_adj else w_half_m_norm
                    curr_diag_len = diag_len_snagi if is_snagi_adj else diag_len_norm
                    curr_angle_deg = angle_m_snagi_deg if is_snagi_adj else angle_m_norm_deg
                    curr_cut_angle = cut_angle_snagi if is_snagi_adj else cut_angle_norm
                    
                    start_tip_m = px + r_left + dx_m
                    end_tip_m = curx - r_right - dx_m
                    
                    if j % 2 != 0:
                        draw_diag_poly(ax2, start_tip_m, end_tip_m, y_inner_bot_m, y_inner_top_m, w_half, True, zorder=3, facecolor='yellow')
                        ax2.text((start_tip_m + end_tip_m)/2, H_truss_cm/2, f"{round(curr_diag_len,1)}cm\n({int(round(curr_cut_angle, 0))}°)", ha='center', va='center', fontsize=18, color='blue', rotation=curr_angle_deg) 
                    else:
                        draw_diag_poly(ax2, start_tip_m, end_tip_m, y_inner_bot_m, y_inner_top_m, w_half, False, zorder=3, facecolor='yellow')
                        ax2.text((start_tip_m + end_tip_m)/2, H_truss_cm/2, f"{round(curr_diag_len,1)}cm\n({int(round(curr_cut_angle, 0))}°)", ha='center', va='center', fontsize=18, color='blue', rotation=-curr_angle_deg) 

        for i in range(n_sec_m):
            span_center = i * gap_m + (gap_m / 2)
            if n_sec_m == 1:
                wall_len = L_wall_single
                w_label = "단일"
            elif i == 0 or i == n_sec_m - 1:
                wall_len = L_wall_end
                w_label = "양끝"
            else:
                wall_len = L_wall_mid
                w_label = "중간"
            
            ax2.text(span_center, -65, f"▼ 벽사다리({w_label}) 가로 절단(스나기 {wall_snagi_mm}mm): {round(wall_len, 1)}cm", 
                     ha='center', color='darkgreen', weight='bold', fontsize=22) 

        ax2.text(-gap_m*0.1, H_truss_cm/2, f"다대 실절단\n{round(actual_main_v_len,1)}cm", va='center', ha='right', color='black', fontsize=24) 
        ax2.set_title(f"2. 메인사다리 상세 (높이: {H_truss_cm}cm, 스나기 30cm 연장)", fontsize=36, fontweight='bold', pad=30) 
        ax2.set_xlim(-L_cm*0.08, L_cm*1.08)
        ax2.set_ylim(-90, H_truss_cm + 100)
        ax2.axis('off'); ax2.set_aspect('equal')

        y_top_center_r = H_ridge_cm - (t_ridge_main_cm / 2)
        y_bot_center_r = 0 + (t_ridge_main_cm / 2)
        y_inner_top_r = H_ridge_cm - t_ridge_main_cm
        y_inner_bot_r = t_ridge_main_cm

        ax3.annotate('', xy=(0, H_ridge_cm + 65), xytext=(L_cm, H_ridge_cm + 65), arrowprops=dict(arrowstyle='<->', color='red', lw=2))
        ax3.text(L_cm/2, H_ridge_cm + 70, f"총기장: {L_cm}cm", ha='center', color='red', weight='bold', fontsize=28) 

        for i in range(n_sec_m + 1):
            cx = i * gap_m
            if i == 0:
                snagi_cx = ridge_deduct_cm / 2  
            elif i == n_sec_m:
                snagi_cx = L_cm - (ridge_deduct_cm / 2) 
            else:
                snagi_cx = cx  
            draw_pipe(ax3, snagi_cx, -30, snagi_cx, H_ridge_cm, ridge_deduct_cm, zorder=5, facecolor='purple')

            if i < n_sec_m:
                cx_start = i * gap_m
                v_centers = []
                
                if n_sec_m == 1:
                    x_min = ridge_deduct_cm
                    x_max = L_cm - ridge_deduct_cm
                    v_centers = [x_min, cx_start + 1*sub_gap, cx_start + 2*sub_gap, cx_start + 3*sub_gap, x_max]
                    label = "[단일세트]"
                elif i == 0:
                    x_min = ridge_deduct_cm
                    x_max = gap_m - ridge_deduct_cm / 2
                    v_centers = [x_min, cx_start + 1*sub_gap, cx_start + 2*sub_gap, cx_start + 3*sub_gap, x_max]
                    label = "[양끝세트]"
                elif i == n_sec_m - 1:
                    x_min = cx_start + ridge_deduct_cm / 2
                    x_max = L_cm - ridge_deduct_cm
                    v_centers = [x_min, cx_start + 1*sub_gap, cx_start + 2*sub_gap, cx_start + 3*sub_gap, x_max]
                    label = "[양끝세트]"
                else:
                    x_min = cx_start + ridge_deduct_cm / 2
                    x_max = (i+1)*gap_m - ridge_deduct_cm / 2
                    v_centers = [x_min, cx_start + 1*sub_gap, cx_start + 2*sub_gap, cx_start + 3*sub_gap, x_max]
                    label = "[중간세트]"

                L_set = x_max - x_min
                ax3.text(cx_start + gap_m/2, H_ridge_cm + 25, f"구분간격(벽사다리 스팬): {round(gap_m,1)}cm", ha='center', color='black', weight='bold', fontsize=24) 

                draw_pipe(ax3, x_min, y_top_center_r, x_max, y_top_center_r, t_ridge_main_cm, zorder=3, facecolor='#C0C0C0')
                draw_pipe(ax3, x_min, y_bot_center_r, x_max, y_bot_center_r, t_ridge_main_cm, zorder=3, facecolor='#C0C0C0')

                for vx in v_centers:
                    draw_pipe(ax3, vx, y_inner_bot_r, vx, y_inner_top_r, t_ridge_v_cm, zorder=4, facecolor='#000080')

                for j in range(4):
                    span_dist = v_centers[j+1] - v_centers[j]
                    
                    is_left_snagi = (j == 0)
                    is_right_snagi = (j == 3)
                    l_r = ridge_deduct_cm/2 if is_left_snagi else t_ridge_v_cm/2
                    r_r = ridge_deduct_cm/2 if is_right_snagi else t_ridge_v_cm/2
                    
                    diag_len, cut_angle, angle_rad, w_half_r = calc_diag(span_dist, actual_ridge_v_len, l_r, r_r, t_ridge_diag_cm)
                    
                    start_tip_r = v_centers[j] + l_r + dx_m
                    end_tip_r = v_centers[j+1] - r_r - dx_m

                    mid_span_x = (v_centers[j] + v_centers[j+1]) / 2
                    
                    txt_color = 'gray' if abs(span_dist - sub_gap) < 0.1 else 'red'
                    span_label = f"자간:\n{round(span_dist,1)}cm"
                    ax3.text(mid_span_x, -H_ridge_cm*0.15, span_label, ha='center', fontsize=18, color=txt_color, weight='bold') 

                    if j % 2 == 0:
                        draw_diag_poly(ax3, start_tip_r, end_tip_r, y_inner_bot_r, y_inner_top_r, w_half_r, True, zorder=3, facecolor='yellow')
                        ax3.text((start_tip_r + end_tip_r)/2, H_ridge_cm/2, f"{round(diag_len,1)}\n({int(round(cut_angle, 0))}°)", ha='center', va='center', fontsize=16, color='blue', rotation=angle_rad) 
                    else:
                        draw_diag_poly(ax3, start_tip_r, end_tip_r, y_inner_bot_r, y_inner_top_r, w_half_r, False, zorder=3, facecolor='yellow')
                        ax3.text((start_tip_r + end_tip_r)/2, H_ridge_cm/2, f"{round(diag_len,1)}\n({int(round(cut_angle, 0))}°)", ha='center', va='center', fontsize=16, color='blue', rotation=-angle_rad) 

                ax3.annotate('', xy=(x_min, -H_ridge_cm*0.45), xytext=(x_max, -H_ridge_cm*0.45), arrowprops=dict(arrowstyle='<->', color='red', lw=1.5))
                ax3.text((x_min + x_max)/2, -H_ridge_cm*0.6, f"{label} 모듈 총길이: {round(L_set,1)}cm", ha='center', fontsize=22, color='red', weight='bold') 
                ax3.text((x_min + x_max)/2, -85, f"▼ 용마루({label}) 가로 절단(공제 {ridge_deduct_mm}mm): {round(L_set, 1)}cm", ha='center', color='darkgreen', weight='bold', fontsize=22) 

        ax3.text(-gap_m*0.1, H_ridge_cm/2, f"다대 실절단\n{round(actual_ridge_v_len,1)}cm", va='center', ha='right', color='black', fontsize=24) 
        
        ax3.set_title(f"3. 용마루 전체 조립도 (높이: {H_ridge_cm}cm, 상하현재 외경: {p_ridge_main}mm)", fontsize=36, fontweight='bold', pad=30) 
        ax3.set_xlim(-L_cm*0.08, L_cm*1.08)
        ax3.set_ylim(-110, H_ridge_cm + 110)
        ax3.axis('off'); ax3.set_aspect('equal')

        plt.tight_layout()
        pdf.savefig(fig)
        
    pdf_buffer.seek(0)
    
    with open(excel_file, "rb") as f:
        excel_bytes = f.read()
        
    return fig, pdf_buffer, excel_bytes, pdf_file_name, excel_file


# ==============================================================================
# Streamlit UI
# ==============================================================================
st.title("🏢 하나천막기업 - 자재 산출 및 도면 생성 시스템")
st.markdown("---")

app_mode = st.sidebar.selectbox("작업 선택", ["1. 맞춤형 트러스 생성기", "2. 벽사다리/보강사다리 통합 산출"])

if app_mode == "1. 맞춤형 트러스 생성기":
    st.header("1. 맞춤형 트러스 생성기")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.subheader("도면 설정")
        type_options = {
            "대칭삼각(일반)": "1", "아치형(일반)": "2", "반삼각(일반)": "3",
            "서브형_삼각": "4", "서브형_아치": "5", "서브형_반삼각": "6",
            "밑더블_삼각": "7", "밑더블_아치": "8", "밑더블_반삼각": "9"
        }
        t_name = st.selectbox("트러스 형태 선택", list(type_options.keys()), index=6)
        type_choice = type_options[t_name]
        
        span_cm = st.number_input("전체 스판(cm)", value=1200.0)
        divs = st.number_input("등분 수(다대 개수 결정)", value=34, step=1)
        
        is_sub_type = type_choice in ["4", "5", "6"]
        has_tie = type_choice in ["4", "5"]
        is_double_bot = type_choice in ["7", "8", "9"]
        
        if is_sub_type and type_choice == "5":
            h_outer_cm = st.number_input("양쪽 끝단(시작) 높이(cm)", value=51.0)
        elif is_sub_type:
            h_outer_cm = st.number_input("상하 일정 수직 폭(깊이)(cm)", value=80.0)
        elif is_double_bot:
            h_outer_cm = st.number_input("밑더블 양끝 외경 시작높이(cm)", value=80.0)
        else:
            h_outer_cm = st.number_input("끝단(시작) 높이(cm)", value=51.0)
            
        h_center_cm = st.number_input("최고점 상단 높이(cm)", value=250.0)
        
        h_tie_cm = 0.0
        if has_tie:
            h_tie_cm = st.number_input("수평재(수평선) 바닥기준 높이(cm)", value=150.0)
            
    with col2:
        st.subheader("파이프 규격(mm)")
        m_od = st.number_input("상/하현부 및 수평 파이프 지름(mm)", value=59.9)
        v_od = st.number_input("다대(일반) 지름(mm)", value=38.1)
        r_od = st.number_input("용마루(중앙) 지름(mm)", value=59.9)
        d_od = st.number_input("살대(대각) 지름(mm)", value=31.8)
        offset_mm = st.number_input("살대 이격 거리(mm)", value=20.0)

    if st.button("도면 및 재단표 생성", type="primary"):
        params = {
            't_name': t_name, 'type_choice': type_choice, 'span_cm': span_cm, 'divs': divs,
            'h_outer_cm': h_outer_cm, 'h_center_cm': h_center_cm, 'h_tie_cm': h_tie_cm,
            'm_od': m_od, 'v_od': v_od, 'r_od': r_od, 'd_od': d_od, 'offset_mm': offset_mm
        }
        
        with st.spinner("도면 및 재단표를 생성 중입니다..."):
            fig, pdf_buffer, excel_bytes, pdf_name, excel_name = generate_custom_truss(params)
            
            st.success("✅ 생성 완료!")
            
            st.pyplot(fig)
            plt.close(fig) 
            
            d_col1, d_col2 = st.columns(2)
            with d_col1:
                st.download_button(label="📥 PDF 도면 다운로드", data=pdf_buffer, file_name=pdf_name, mime="application/pdf")
            with d_col2:
                st.download_button(label="📥 엑셀 재단표 다운로드", data=excel_bytes, file_name=excel_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


elif app_mode == "2. 벽사다리/보강사다리 통합 산출":
    st.header("2. 벽사다리 및 보강사다리 통합 산출 시스템")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.subheader("1. 기본 치수 (중심선 기준)")
        L_cm = st.number_input("전체 총기장(cm)", value=2000.0)
        W_cm = st.number_input("보강사다리 폭(cm)", value=70.0)
        H_truss_cm = st.number_input("메인사다리 높이(폭)(cm)", value=70.0)
        total_sets_sub = st.number_input("보강사다리 총 제작 수량(세트)", value=1, step=1)
        total_sets_main = st.number_input("메인사다리 총 제작 수량(세트)", value=1, step=1)
        offset_mm = st.number_input("살대 다대 꼭지점 이격 거리(mm)", value=10.0)
        
    with col2:
        st.subheader("2. 용마루 & 벽사다리")
        H_ridge_cm = st.number_input("용마루 폭(높이)(cm)", value=70.0)
        ridge_deduct_mm = st.number_input("용마루 공제 기준 사이즈(mm)", value=59.9)
        total_sets_ridge = st.number_input("용마루 전체 라인 제작 수량(세트)", value=1, step=1)
        wall_snagi_mm = st.number_input("벽사다리 스나기 사이즈(mm)", value=89.1)
        
    with col3:
        st.subheader("3. 파이프 규격(mm)")
        p_sub_main = st.number_input("보강사다리 상하현재 지름(mm)", value=38.1)
        p_sub_sub = st.number_input("보강사다리 수직/사재 지름(mm)", value=31.8)
        p_main_main = st.number_input("메인사다리 상하현재 지름(mm)", value=42.2)
        p_main_snagi = st.number_input("메인사다리 스나기 지름(mm)", value=89.1)
        p_main_v = st.number_input("메인사다리 수직다대 지름(mm)", value=38.1)
        p_main_diag = st.number_input("메인사다리 사다리살대 지름(mm)", value=31.8)
        p_ridge_main = st.number_input("용마루 상하현재 지름(mm)", value=42.2)
        p_ridge_v = st.number_input("용마루 수직다대 지름(mm)", value=38.1)
        p_ridge_diag = st.number_input("용마루 사다리살대 지름(mm)", value=31.8)
        
    if st.button("산출표 및 도면 생성", type="primary"):
        params = {
            'L_cm': L_cm, 'W_cm': W_cm, 'H_truss_cm': H_truss_cm, 
            'total_sets_sub': total_sets_sub, 'total_sets_main': total_sets_main,
            'offset_mm': offset_mm, 'H_ridge_cm': H_ridge_cm, 
            'ridge_deduct_mm': ridge_deduct_mm, 'total_sets_ridge': total_sets_ridge,
            'wall_snagi_mm': wall_snagi_mm, 'p_sub_main': p_sub_main, 'p_sub_sub': p_sub_sub,
            'p_main_main': p_main_main, 'p_main_snagi': p_main_snagi, 'p_main_v': p_main_v,
            'p_main_diag': p_main_diag, 'p_ridge_main': p_ridge_main, 'p_ridge_v': p_ridge_v, 'p_ridge_diag': p_ridge_diag
        }
        
        with st.spinner("산출표와 도면을 생성 중입니다..."):
            fig, pdf_buffer, excel_bytes, pdf_name, excel_name = run_ladder_system(params)
            
            st.success("✅ 생성 완료!")
            
            st.pyplot(fig)
            plt.close(fig)
            
            d_col1, d_col2 = st.columns(2)
            with d_col1:
                st.download_button(label="📥 PDF 상세도면 다운로드", data=pdf_buffer, file_name=pdf_name, mime="application/pdf")
            with d_col2:
                st.download_button(label="📥 엑셀 종합산출표 다운로드", data=excel_bytes, file_name=excel_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
